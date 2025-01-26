import os
from typing import LiteralString
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, PatternFill, NamedStyle


def get_files(dir_name):
    list_files = []
    for root, dirs, files in os.walk(dir_name):
        list_files: list[LiteralString | str | bytes] = [os.path.join(root, file) for file in files if file.endswith('csv')]
    return list_files


def get_data_frame(file):
    df = pd.read_csv(file, encoding='cp1251', delimiter=';', quotechar='"')
    # Удаляем символы ="..." из заголовков и ячеек
    df.columns = df.columns.str.replace(r'="(.*?)"', r'\1', regex=True)
    df = df.replace(r'="(.*?)"', r'\1', regex=True)
    return df


def convert_to_number(df):
    # Преобразование столбцов в числовой формат
    columns_to_convert = ['ФВ до', 'ФВ после', 'ФВ без инд.до', 'ФВ без инд.после',
                          'СП до', 'СП после', 'СП без инд. до', 'СП без инд. после',
                          'ГП1 до', 'ГП1 после', 'ГП2 до', 'ГП2 после',
                          'Доля до', 'Доля после']
    df[columns_to_convert] = df[columns_to_convert].apply(pd.to_numeric, errors='coerce')
    return df


def add_columns(df):
    # Добавляем столбцы
    df['Разница ФВ'] = ''
    df['% увел.ФВ'] = ''
    df['Разница ФВ без инд'] = ''
    df['Не кор.инд ФВ'] = ''
    df['Разница СП'] = ''
    df['% увел.СП'] = ''
    df['Разница СП без инд'] = ''
    df['Не кор.инд Стр'] = ''
    df['Разница ГП1'] = ''
    df['Разница ГП2'] = ''
    df['% увелич.доля'] = ''

    # Размещение столбцов
    cols = df.columns.tolist()
    cols.insert(cols.index('ФВ после') + 1, cols.pop(cols.index('Разница ФВ')))
    cols.insert(cols.index('Разница ФВ') + 1, cols.pop(cols.index('% увел.ФВ')))
    cols.insert(cols.index('ФВ без инд.после') + 1, cols.pop(cols.index('Разница ФВ без инд')))
    cols.insert(cols.index('Разница ФВ без инд') + 1, cols.pop(cols.index('Не кор.инд ФВ')))
    cols.insert(cols.index('СП после') + 1, cols.pop(cols.index('Разница СП')))
    cols.insert(cols.index('Разница СП') + 1, cols.pop(cols.index('% увел.СП')))
    cols.insert(cols.index('СП без инд. после') + 1, cols.pop(cols.index('Разница СП без инд')))
    cols.insert(cols.index('Разница СП без инд') + 1, cols.pop(cols.index('Не кор.инд Стр')))
    cols.insert(cols.index('ГП1 после') + 1, cols.pop(cols.index('Разница ГП1')))
    cols.insert(cols.index('ГП2 после') + 1, cols.pop(cols.index('Разница ГП2')))
    cols.insert(cols.index('Доля после') + 1, cols.pop(cols.index('% увелич.доля')))
    df = df[cols]
    return df


def add_formulas(df, file, dir_name_result):
    file_name = os.path.basename(file)
    new_file = os.path.join(dir_name_result, file_name)
    new_file = new_file.replace('csv', 'xlsx')
    with pd.ExcelWriter(new_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        worksheet = writer.sheets['Sheet1']
        headers = [cell.value for cell in worksheet[1]]

        for col_num, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_num, value=header)

            # Устанавливаем выравнивание с переносом текста
            cell.alignment = Alignment(wrap_text=True)

        # Закрашиваем заголовки колонок
        header_cell = worksheet['I1']
        header_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                       fill_type="solid")  # Закрашиваем в желтый цвет
        header_cell = worksheet['J1']
        header_cell.fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF",
                                       fill_type="solid")  # Закрашиваем в желтый цвет
        header_cell = worksheet['M1']
        header_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                       fill_type="solid")  # Закрашиваем в желтый цвет
        header_cell = worksheet['N1']
        header_cell.fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF",
                                       fill_type="solid")  # Закрашиваем в желтый цвет
        header_cell = worksheet['Q1']
        header_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                       fill_type="solid")  # Закрашиваем в желтый цвет
        header_cell = worksheet['R1']
        header_cell.fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF",
                                       fill_type="solid")  # Закрашиваем в желтый цвет
        header_cell = worksheet['U1']
        header_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                       fill_type="solid")  # Закрашиваем в желтый цвет
        header_cell = worksheet['V1']
        header_cell.fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF",
                                       fill_type="solid")  # Закрашиваем в желтый цвет
        header_cell = worksheet['Y1']
        header_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                       fill_type="solid")  # Закрашиваем в желтый цвет
        header_cell = worksheet['AD1']
        header_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                       fill_type="solid")  # Закрашиваем в желтый цвет
        header_cell = worksheet['AK1']
        header_cell.fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF",
                                       fill_type="solid")  # Закрашиваем в желтый цвет

        number_format = NamedStyle(name='number_format', number_format='0.00')  # Установите нужный формат

        for i in range(2, len(df) + 2):
            worksheet[f"I{i}"] = f"=ROUND(G{i}*1.095, 2)-H{i}"
            worksheet[f"I{i}"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                  fill_type="solid")
            worksheet[f"J{i}"] = f"=ROUND((H{i}/G{i}-1)*100, 1)"
            worksheet[f"J{i}"].fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF",
                                                  fill_type="solid")
            worksheet[f"M{i}"] = f"=ROUND(L{i}-K{i}, 2)"
            worksheet[f"M{i}"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                  fill_type="solid")
            worksheet[f"N{i}"] = f"=IF(ROUND(H{i}-G{i}, 2)=M{i}, 0, ROUND(L{i}-H{i}, 2))"
            worksheet[f"N{i}"].style =  number_format # Убедитесь, что стиль числовой
            worksheet[f"N{i}"].fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF",
                                                  fill_type="solid")
            worksheet[f"Q{i}"] = f"=ROUND(O{i}*1.095, 2)-P{i}"
            worksheet[f"Q{i}"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                  fill_type="solid")
            worksheet[f"R{i}"] = f"=ROUND((P{i}/O{i}-1)*100, 1)"
            worksheet[f"R{i}"].fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF",
                                                  fill_type="solid")
            worksheet[f"U{i}"] = f"=ROUND(T{i}-S{i}, 2)"
            worksheet[f"U{i}"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                  fill_type="solid")
            worksheet[f"V{i}"] = f"=IF(ROUND(P{i}-O{i}, 2)=U{i}, 0, ROUND(T{i}-P{i}, 2))"
            worksheet[f"V{i}"].fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF",
                                                  fill_type="solid")
            worksheet[f"Y{i}"] = f"=X{i}-W{i}"
            worksheet[f"Y{i}"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                  fill_type="solid")
            worksheet[f"AD{i}"] = f"=AC{i}-AB{i}"
            worksheet[f"AD{i}"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                  fill_type="solid")
            worksheet[f"AK{i}"] = f"=ROUND((AJ{i}/AI{i}-1)*100, 1)"
            worksheet[f"AK{i}"].fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF",
                                                  fill_type="solid")


def main(dir_name, dir_name_result):
    files = get_files(dir_name)
    print('Start')
    for file in files:
        df = get_data_frame(file)
        df = convert_to_number(df)
        df = add_columns(df)
        add_formulas(df, file, dir_name_result)
        print(file)
    print('End')

if __name__ == '__main__':
    folder = r"D:\Python\SFR\data"
    folder_result = r"D:\Python\SFR\output"

    main(folder, folder)
