import os
from typing import LiteralString
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, PatternFill, NamedStyle


def get_files(dir_name):
    list_files = []
    for root, dirs, files in os.walk(dir_name):
        list_files: list[LiteralString | str | bytes] = [os.path.join(root, file) for file in files if file.endswith('xlsx')]
    return list_files


def get_data_frame(file):
    df = pd.read_excel(file, engine='openpyxl')
    return df


def filtration(df):
    filtered_df = df[(df['Не кор.инд ФВ'] != 0) | ((df['Не кор.инд Стр'] != 0) & (df['Не кор.инд Стр'] != -0.01))]
    return filtered_df


def connect_df(result, df):
    result = pd.concat([result, df], ignore_index=True)
    return result


def add_formulas(df, file, dir_name_result):
    file_name = os.path.basename(file)
    new_file = os.path.join(dir_name_result, file_name)
    new_file = new_file.replace('csv', 'xlsx')
    with pd.ExcelWriter(new_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        worksheet = writer.sheets['Sheet1']
        headers = [cell.value for cell in worksheet[1]]

        for col_num, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_num, value=header)

            # Устанавливаем выравнивание с переносом текста
            cell.alignment = Alignment(wrap_text=True)

        # Закрашиваем заголовки колонок
        header_cell = worksheet['I1']
        header_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                       fill_type="solid")  # Закрашиваем в желтый цвет
        header_cell = worksheet['J1']
        header_cell.fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF",
                                       fill_type="solid")  # Закрашиваем в желтый цвет
        header_cell = worksheet['M1']
        header_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                       fill_type="solid")  # Закрашиваем в желтый цвет
        header_cell = worksheet['N1']
        header_cell.fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF",
                                       fill_type="solid")  # Закрашиваем в желтый цвет
        header_cell = worksheet['Q1']
        header_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                       fill_type="solid")  # Закрашиваем в желтый цвет
        header_cell = worksheet['R1']
        header_cell.fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF",
                                       fill_type="solid")  # Закрашиваем в желтый цвет
        header_cell = worksheet['U1']
        header_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                       fill_type="solid")  # Закрашиваем в желтый цвет
        header_cell = worksheet['V1']
        header_cell.fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF",
                                       fill_type="solid")  # Закрашиваем в желтый цвет
        header_cell = worksheet['Y1']
        header_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                       fill_type="solid")  # Закрашиваем в желтый цвет
        header_cell = worksheet['AD1']
        header_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                       fill_type="solid")  # Закрашиваем в желтый цвет
        header_cell = worksheet['AK1']
        header_cell.fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF",
                                       fill_type="solid")  # Закрашиваем в желтый цвет
        number_format = NamedStyle(name='number_format', number_format='0.00')  # Установите нужный формат

        for i in range(2, len(df) + 2):
            worksheet[f"I{i}"] = f"=ROUND(G{i}*1.095, 2)-H{i}"
            worksheet[f"I{i}"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                  fill_type="solid")
            worksheet[f"J{i}"] = f"=ROUND((H{i}/G{i}-1)*100, 1)"
            worksheet[f"J{i}"].fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF",
                                                  fill_type="solid")
            worksheet[f"M{i}"] = f"=ROUND(L{i}-K{i}, 2)"
            worksheet[f"M{i}"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                  fill_type="solid")
            worksheet[f"N{i}"] = f"=IF(ROUND(H{i}-G{i}, 2)=M{i}, 0, ROUND(L{i}-H{i}, 2))"
            worksheet[f"N{i}"].style =  number_format # Убедитесь, что стиль числовой
            worksheet[f"N{i}"].fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF",
                                                  fill_type="solid")
            worksheet[f"Q{i}"] = f"=ROUND(O{i}*1.095, 2)-P{i}"
            worksheet[f"Q{i}"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                  fill_type="solid")
            worksheet[f"R{i}"] = f"=ROUND((P{i}/O{i}-1)*100, 1)"
            worksheet[f"R{i}"].fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF",
                                                  fill_type="solid")
            worksheet[f"U{i}"] = f"=ROUND(T{i}-S{i}, 2)"
            worksheet[f"U{i}"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                  fill_type="solid")
            worksheet[f"V{i}"] = f"=IF(ROUND(P{i}-O{i}, 2)=U{i}, 0, ROUND(T{i}-P{i}, 2))"
            worksheet[f"V{i}"].fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF",
                                                  fill_type="solid")
            worksheet[f"Y{i}"] = f"=X{i}-W{i}"
            worksheet[f"Y{i}"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                  fill_type="solid")
            worksheet[f"AD{i}"] = f"=AC{i}-AB{i}"
            worksheet[f"AD{i}"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                  fill_type="solid")
            worksheet[f"AK{i}"] = f"=ROUND((AJ{i}/AI{i}-1)*100, 1)"
            worksheet[f"AK{i}"].fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF",
                                                  fill_type="solid")

def main(dir_name, dir_name_result):
    files = get_files(dir_name)
    print('Start')
    result = pd.DataFrame()
    for file in files:
        df = get_data_frame(file)
        df = filtration(df)
        result = connect_df(result, df)

    add_formulas(result, 'all_plus.xlsx', 'output')
    print(result)
    print('End')

if __name__ == '__main__':
    folder = r"D:\Python\SFR\data"
    folder_result = r"D:\Python\SFR\output"

    main(folder, folder)